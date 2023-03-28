import argparse
from openpyxl import load_workbook
from collections import namedtuple
from excel2tex.Table import Table

def run():
    parser = argparse.ArgumentParser(
            description='Convert excel table to latex table.',
            formatter_class=argparse.RawTextHelpFormatter,
            epilog="""Note: The height of every merged cell must not be less than the number of lines in your text.
"""
            )
    parser.add_argument('-s', default='table.xlsx', dest='source', help='source file (default: %(default)s)')
    parser.add_argument('-o', default='table.tex', dest='target', help='target file (default: %(default)s)')
    parser.add_argument('-p', default='setting.tex', dest='setting', help='setting file (default: %(default)s)')
    parser.add_argument('-w', default='\\linewidth', dest='width', help='table width (default: %(default)s)')
    parser.add_argument('--sig', default='utf-8', dest='encoding',
            nargs='?',
            const='utf-8-sig',
            help='set file encoding to utf-8-sig, only use when there is mess code.')
    parser.add_argument('-m', '--math', type=bool, default=False, dest='math', help='enabel inline math', const=True, nargs='?')
    parser.add_argument('-e', '--excel-format', type=bool, default=False, dest='excel_format', help='enabel inline math', const=True, nargs='?')
    args = parser.parse_args()
    wb = load_workbook(args.source)
    ws = wb.active
    t = Table(ws, args)
    with open(args.target, 'w', encoding=args.encoding) as f:
        f.write(t.tex)

def convert_table(source, target="", setting="setting.tex", width="\\linewidth", encoding="utf-8", math=False, excel_format=False):
    excel2tex_args = namedtuple('excel2tex_args', ['source', 'target', 'setting', 'width', 'encoding', 'math', 'excel_format'])
    wb = load_workbook(source)
    ws = wb.active
    args = excel2tex_args(source, target, setting, width, encoding, math, excel_format)
    t = Table(ws, args)
    if target:
        with open(target, 'w', encoding=encoding) as f:
                f.write(t.tex)
    else:
        return t.tex

if __name__ == '__main__':
    run()